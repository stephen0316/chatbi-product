from __future__ import annotations

import argparse
import json
from collections import Counter, defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


TARGET_REVENUE_SHEET = "收入及直接成本明细表"
PRODUCT_LIST_SHEETS = ("自研类", "引入类")
ACTIVE_STATUSES_FOR_RULES_1_TO_3 = {"已上市", "已入库"}
REVENUE_START_MONTH = "2024-06"
REVENUE_END_MONTH = "2026-05"
ERROR_LITERALS = {"#N/A", "#VALUE!", "#REF!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!"}


def subtract_years(value: date, years: int) -> date:
    try:
        return value.replace(year=value.year - years)
    except ValueError:
        return value.replace(year=value.year - years, day=28)


AS_OF_DATE = date.today()
OLDER_THAN_TWO_YEARS_BEFORE = subtract_years(AS_OF_DATE, 2)
RULE4_APPROVAL_BEFORE = subtract_years(AS_OF_DATE, 1)
OUTPUT_HEADERS = [
    "产品编码",
    "产品名称",
    "产品状态",
    "创建时间",
    "部门",
    "产品收入",
    "产品毛利",
    "退市类型",
    "理由",
]
RULE_TEXT = {
    "1": "已上市/已入库但未出现在收入及直接成本明细表",
    "2": "创建时间超过2年或为空，且2年内产品收入为0",
    "3": "创建时间超过2年或为空，且2年内产品毛利≤0",
    "4": "产品状态为退市中，且退市审批完成时间超过1年",
}

PRODUCT_COLUMN_ALIASES = {
    "code": ["产品编码", "产品编号", "产品代码"],
    "name": ["产品名称", "产品名"],
    "status": ["产品状态", "状态"],
    "created": ["产品创建时间", "创建时间", "创建日期", "产品创建日期"],
    "department": ["产品所属部门", "所属部门", "部门", "责任部门"],
    "delisting_approval_completed": [
        "退市审批完成时间",
        "退市审批完成日期",
        "审批完成时间",
        "审批完成日期",
        "退市完成时间",
        "退市完成日期",
    ],
}
REVENUE_COLUMN_ALIASES = {
    "code": ["产品编码", "产品编号", "产品代码", "编码"],
    "revenue": ["产品收入", "收入", "收入金额", "金额", "营业收入", "销售收入"],
    "gross_profit": ["产品毛利", "毛利", "毛利额", "毛利润"],
    "month": ["收入月份", "月份", "收入年月", "账期", "会计期间", "日期"],
}
PUBLIC_COLUMN_LABELS = {
    "code": "产品编码",
    "name": "产品名称",
    "status": "产品状态",
    "created": "创建时间",
    "department": "部门",
    "delisting_approval_completed": "退市审批完成时间",
    "revenue": "产品收入",
    "gross_profit": "产品毛利",
    "month": "收入月份",
}


def normalize(value: Any) -> str:
    if value is None or isinstance(value, bool):
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    return "" if text in ERROR_LITERALS else text


def parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = normalize(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def month_key(value: Any) -> str:
    parsed = parse_date(value)
    if parsed:
        return parsed.strftime("%Y-%m")
    return normalize(value).replace("/", "-")[:7]


def required_revenue_months() -> list[str]:
    start_year, start_month = (int(part) for part in REVENUE_START_MONTH.split("-"))
    end_year, end_month = (int(part) for part in REVENUE_END_MONTH.split("-"))
    months: list[str] = []
    year = start_year
    month = start_month
    while (year, month) <= (end_year, end_month):
        months.append(f"{year:04d}-{month:02d}")
        month += 1
        if month == 13:
            year += 1
            month = 1
    return months


def summarize_months(months: set[str]) -> dict[str, Any]:
    valid_months = {month for month in months if len(month) == 7 and month[4] == "-"}
    sorted_months = sorted(valid_months)
    required = required_revenue_months()
    missing = [month for month in required if month not in valid_months]
    return {
        "months": sorted_months,
        "start_month": sorted_months[0] if sorted_months else "",
        "end_month": sorted_months[-1] if sorted_months else "",
        "covers_required_window": not missing,
        "missing_months": missing,
    }


def to_decimal(value: Any) -> Decimal:
    if value is None or value == "" or isinstance(value, bool):
        return Decimal("0")
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, ValueError):
        return Decimal("0")


def decimal_to_number(value: Decimal) -> int | float:
    rounded = value.quantize(Decimal("0.01"))
    return int(rounded) if rounded == rounded.to_integral() else float(rounded)


def first_existing(headers: list[str], names: list[str]) -> int | None:
    for name in names:
        if name in headers:
            return headers.index(name)
    return None


def alias_index(headers: list[str], aliases: list[str]) -> int | None:
    normalized_headers = [normalize(header).replace(" ", "") for header in headers]
    for alias in aliases:
        normalized_alias = alias.replace(" ", "")
        if normalized_alias in normalized_headers:
            return normalized_headers.index(normalized_alias)
    for alias in aliases:
        normalized_alias = alias.replace(" ", "")
        for index, header in enumerate(normalized_headers):
            if normalized_alias and normalized_alias in header:
                return index
    return None


def find_mapping(headers: list[str], aliases: dict[str, list[str]]) -> dict[str, int | None]:
    return {field: alias_index(headers, field_aliases) for field, field_aliases in aliases.items()}


def mapping_score(mapping: dict[str, int | None], required: list[str]) -> int:
    return sum(1 for field in required if mapping.get(field) is not None)


def has_required_mapping(mapping: dict[str, int | None], required: list[str]) -> bool:
    return all(mapping.get(field) is not None for field in required)


def get_cell(row: tuple[Any, ...] | list[Any], index: int | None) -> Any:
    if index is None or index >= len(row):
        return None
    return row[index]


def find_sheet_profile(sheet) -> dict[str, Any]:
    best: dict[str, Any] | None = None
    max_row = min(sheet.max_row, 12)
    max_column = max(sheet.max_column or 0, 120)
    for row_index in range(1, max_row + 1):
        headers = [normalize(sheet.cell(row_index, col).value) for col in range(1, max_column + 1)]
        if not any(headers):
            continue
        product_mapping = find_mapping(headers, PRODUCT_COLUMN_ALIASES)
        revenue_mapping = find_mapping(headers, REVENUE_COLUMN_ALIASES)
        product_score = mapping_score(product_mapping, ["code", "name", "status", "created", "department"])
        revenue_score = mapping_score(revenue_mapping, ["code", "revenue", "gross_profit", "month"])
        score = max(product_score, revenue_score)
        if best is None or score > best["score"]:
            best = {
                "header_row": row_index,
                "headers": headers,
                "product_mapping": product_mapping,
                "revenue_mapping": revenue_mapping,
                "product_score": product_score,
                "revenue_score": revenue_score,
                "score": score,
            }
    if best is None:
        return {
            "header_row": 1,
            "headers": [],
            "product_mapping": {},
            "revenue_mapping": {},
            "product_score": 0,
            "revenue_score": 0,
            "score": 0,
        }
    return best


def serialize_columns(headers: list[str], mapping: dict[str, int | None]) -> dict[str, dict[str, Any] | None]:
    columns: dict[str, dict[str, Any] | None] = {}
    for field, index in mapping.items():
        if index is None:
            columns[field] = None
        else:
            columns[field] = {
                "index": index,
                "header": headers[index] if index < len(headers) else "",
                "label": PUBLIC_COLUMN_LABELS.get(field, field),
            }
    return columns


def data_start_row(sheet_name: str, header_row: int) -> int:
    if sheet_name == "自研类" and header_row == 1:
        return 4
    return header_row + 1


def count_rows_with_code(sheet, start_row: int, code_idx: int | None, max_col: int | None = None) -> int:
    if code_idx is None:
        return 0
    count = 0
    for row in sheet.iter_rows(min_row=start_row, max_col=max_col, values_only=True):
        if normalize(get_cell(row, code_idx)):
            count += 1
    return count


def month_range_for_sheet(sheet, start_row: int, month_idx: int | None, max_col: int | None = None) -> dict[str, Any]:
    months: set[str] = set()
    if month_idx is None:
        return summarize_months(months)
    for row in sheet.iter_rows(min_row=start_row, max_col=max_col, values_only=True):
        month = month_key(get_cell(row, month_idx))
        if month:
            months.add(month)
    return summarize_months(months)


def scan_profile_rows(
    sheet,
    start_row: int,
    product_code_idx: int | None,
    revenue_code_idx: int | None,
    month_idx: int | None,
    max_col: int,
) -> tuple[int, int, dict[str, Any]]:
    product_row_count = 0
    revenue_row_count = 0
    months: set[str] = set()
    for row in sheet.iter_rows(min_row=start_row, max_col=max_col, values_only=True):
        if normalize(get_cell(row, product_code_idx)):
            product_row_count += 1
        if normalize(get_cell(row, revenue_code_idx)):
            revenue_row_count += 1
        month = month_key(get_cell(row, month_idx))
        if month:
            months.add(month)
    return product_row_count, revenue_row_count, summarize_months(months)


def require_index(header_map: dict[str, int], name: str, sheet_name: str) -> int:
    if name not in header_map:
        raise ValueError(f"{sheet_name} 缺少必要字段：{name}")
    return header_map[name]


def read_product_list(product_list_path: Path) -> tuple[dict[str, dict[str, Any]], dict[str, Any]]:
    workbook = load_workbook(product_list_path, read_only=False, data_only=True)
    products: dict[str, dict[str, Any]] = {}
    status_counter: Counter[str] = Counter()
    sheet_counts: dict[str, int] = {}

    for sheet_name in PRODUCT_LIST_SHEETS:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"产品全量列表缺少 sheet：{sheet_name}")
        sheet = workbook[sheet_name]
        headers = [normalize(sheet.cell(1, col).value) for col in range(1, sheet.max_column + 1)]
        header_map = {header: idx for idx, header in enumerate(headers) if header}
        code_idx = require_index(header_map, "产品编码", sheet.title)
        name_idx = require_index(header_map, "产品名称", sheet.title)
        status_idx = require_index(header_map, "产品状态", sheet.title)
        created_idx = require_index(header_map, "产品创建时间", sheet.title)
        department_idx = require_index(header_map, "产品所属部门", sheet.title)
        approval_idx = first_existing(
            headers,
            PRODUCT_COLUMN_ALIASES["delisting_approval_completed"],
        )
        start_row = 4 if sheet.title == "自研类" else 2
        sheet_count = 0

        for row_index in range(start_row, sheet.max_row + 1):
            row = [sheet.cell(row_index, col).value for col in range(1, sheet.max_column + 1)]
            code = normalize(row[code_idx])
            name = normalize(row[name_idx])
            if not code and not name:
                continue
            if not code:
                raise ValueError(f"{sheet.title} 第 {row_index} 行缺少产品编码，无法按产品编码匹配")
            if code in products:
                raise ValueError(f"全量产品清单存在重复产品编码：{code}")
            status = normalize(row[status_idx])
            approval_completed = get_cell(row, approval_idx)
            products[code] = {
                "产品编码": code,
                "产品名称": name,
                "产品状态": status,
                "创建时间": normalize(row[created_idx]),
                "创建日期": parse_date(row[created_idx]),
                "部门": normalize(row[department_idx]),
                "退市审批完成时间": normalize(approval_completed),
                "退市审批完成日期": parse_date(approval_completed),
                "来源sheet": sheet.title,
                "来源行号": row_index,
            }
            status_counter[status] += 1
            sheet_count += 1
        sheet_counts[sheet.title] = sheet_count

    return products, {
        "product_count": len(products),
        "status_counts": dict(status_counter),
        "sheet_counts": sheet_counts,
    }


def inspect_workbooks(paths: list[Path]) -> dict[str, Any]:
    files: list[dict[str, Any]] = []
    product_candidates: list[dict[str, Any]] = []
    revenue_candidates: list[dict[str, Any]] = []
    warnings: list[str] = []

    for file_index, workbook_path in enumerate(paths):
        file_id = f"file_{file_index + 1}"
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        if any(sheet.max_row > 1 and sheet.max_column <= 1 for sheet in workbook.worksheets):
            workbook.close()
            workbook = load_workbook(workbook_path, read_only=False, data_only=True)
        sheets: list[dict[str, Any]] = []
        file_product_score = 0
        file_revenue_score = 0
        file_product_candidates = 0
        for sheet in workbook.worksheets:
            profile = find_sheet_profile(sheet)
            start_row = data_start_row(sheet.title, profile["header_row"])
            product_columns = serialize_columns(profile["headers"], profile["product_mapping"])
            revenue_columns = serialize_columns(profile["headers"], profile["revenue_mapping"])
            product_row_count = 0
            revenue_row_count = 0
            month_summary = summarize_months(set())
            if profile["product_score"] >= 4 or profile["revenue_score"] >= 4:
                scan_indexes = [
                    profile["product_mapping"].get("code"),
                    profile["revenue_mapping"].get("code"),
                    profile["revenue_mapping"].get("month"),
                ]
                max_col = max([idx for idx in scan_indexes if idx is not None] or [0]) + 1
                product_row_count, revenue_row_count, month_summary = scan_profile_rows(
                    sheet,
                    start_row,
                    profile["product_mapping"].get("code"),
                    profile["revenue_mapping"].get("code"),
                    profile["revenue_mapping"].get("month"),
                    max_col,
                )
            file_product_score += profile["product_score"]
            file_revenue_score += profile["revenue_score"]
            sheet_info = {
                "name": sheet.title,
                "header_row": profile["header_row"],
                "data_start_row": start_row,
                "headers": [header for header in profile["headers"] if header],
                "product_score": profile["product_score"],
                "revenue_score": profile["revenue_score"],
                "product_columns": product_columns,
                "revenue_columns": revenue_columns,
                "product_row_count": product_row_count,
                "revenue_row_count": revenue_row_count,
                "month_summary": month_summary,
            }
            sheets.append(sheet_info)
            is_product_candidate = has_required_mapping(
                profile["product_mapping"],
                ["code", "name", "status", "created", "department"],
            )
            if is_product_candidate:
                file_product_candidates += 1
                product_candidates.append(
                    {
                        "file_id": file_id,
                        "sheet": sheet.title,
                        "header_row": profile["header_row"],
                        "data_start_row": start_row,
                        "columns": product_columns,
                        "row_count": product_row_count,
                    }
                )
            if profile["revenue_score"] >= 4 and revenue_row_count >= 50:
                revenue_candidates.append(
                    {
                        "file_id": file_id,
                        "sheet": sheet.title,
                        "header_row": profile["header_row"],
                        "data_start_row": start_row,
                        "columns": revenue_columns,
                        "row_count": revenue_row_count,
                        "month_summary": month_summary,
                    }
                )

        role = "unknown"
        confidence = 0.35
        reason = "未识别到足够字段"
        if file_product_candidates > 0 and file_product_score >= file_revenue_score:
            role = "product_list"
            confidence = 0.9 if {"自研类", "引入类"} & set(workbook.sheetnames) else 0.78
            reason = "识别到产品编码、名称、状态、创建时间、部门等产品字段"
        elif file_revenue_score >= 4 and file_revenue_score > file_product_score:
            role = "revenue_detail"
            confidence = 0.86
            reason = "识别到产品编码、收入、毛利、月份等经营字段"
        files.append(
            {
                "id": file_id,
                "name": workbook_path.name,
                "path": str(workbook_path),
                "role": role,
                "confidence": confidence,
                "reason": reason,
                "sheets": sheets,
            }
        )

    product_file_ids = [candidate["file_id"] for candidate in product_candidates]
    selected_product_file_id = product_file_ids[0] if product_file_ids else ""
    selected_product_sheets = [
        candidate for candidate in product_candidates if candidate["file_id"] == selected_product_file_id
    ]
    if selected_product_file_id:
        product_file = next((file for file in files if file["id"] == selected_product_file_id), None)
        existing_standard_sheets = {
            sheet["name"] for sheet in (product_file or {}).get("sheets", []) if sheet["name"] in PRODUCT_LIST_SHEETS
        }
        selected_standard_sheets = {item["sheet"] for item in selected_product_sheets}
        missing_standard_sheets = sorted(existing_standard_sheets - selected_standard_sheets)
        if missing_standard_sheets:
            warnings.append(
                f"产品全量列表的 {'、'.join(missing_standard_sheets)} 缺少产品编码、产品名称、产品状态、创建时间、部门中的必要字段，请修正后重新预检。"
            )
            if product_file is not None:
                product_file["role"] = "unknown"
                product_file["confidence"] = 0.35
                product_file["reason"] = "产品全量列表关键字段不完整"
            selected_product_file_id = ""
            selected_product_sheets = []
    selected_revenue_file_ids = sorted({candidate["file_id"] for candidate in revenue_candidates})
    selected_revenue_sheets = revenue_candidates

    if not selected_product_file_id:
        warnings.append("未能自动识别产品全量列表，请检查是否包含产品编码、名称、状态、创建时间、部门字段。")
    if not selected_revenue_file_ids:
        warnings.append("未能自动识别收入成本明细表，请检查是否包含产品编码、收入、毛利、月份字段。")
    revenue_month_summary = summarize_months(
        {
            month
            for item in selected_revenue_sheets
            for month in item.get("month_summary", {}).get("months", [])
        }
    )
    if selected_revenue_file_ids and not revenue_month_summary["covers_required_window"]:
        start = revenue_month_summary["start_month"] or "未知"
        end = revenue_month_summary["end_month"] or "未知"
        missing = "、".join(revenue_month_summary["missing_months"][:8])
        suffix = "等" if len(revenue_month_summary["missing_months"]) > 8 else ""
        warnings.append(
            f"全部收入成本明细表合并后的月份范围为 {start} 至 {end}，未覆盖完整口径 {REVENUE_START_MONTH} 至 {REVENUE_END_MONTH}，缺少 {missing}{suffix}。"
        )

    return {
        "files": files,
        "selected": {
            "product_file_id": selected_product_file_id,
            "product_sheets": selected_product_sheets,
            "revenue_file_ids": selected_revenue_file_ids,
            "revenue_sheets": selected_revenue_sheets,
            "revenue_month_summary": revenue_month_summary,
        },
        "warnings": warnings,
        "needs_ai": not selected_product_file_id or not selected_revenue_file_ids,
        "needs_confirmation": True,
        "revenue_window": {"start_month": REVENUE_START_MONTH, "end_month": REVENUE_END_MONTH},
    }


def column_index(column: dict[str, Any] | None, label: str, sheet_name: str) -> int:
    if not column or column.get("index") is None:
        raise ValueError(f"{sheet_name} 缺少必要字段：{label}")
    return int(column["index"])


def read_product_list_from_mapping(inspection: dict[str, Any]) -> tuple[dict[str, dict[str, Any]], dict[str, Any]]:
    files_by_id = {item["id"]: item for item in inspection["files"]}
    products: dict[str, dict[str, Any]] = {}
    status_counter: Counter[str] = Counter()
    sheet_counts: dict[str, int] = {}

    for sheet_mapping in inspection["selected"]["product_sheets"]:
        file_info = files_by_id[sheet_mapping["file_id"]]
        workbook = load_workbook(file_info["path"], read_only=False, data_only=True)
        sheet = workbook[sheet_mapping["sheet"]]
        columns = sheet_mapping["columns"]
        code_idx = column_index(columns.get("code"), "产品编码", sheet.title)
        name_idx = column_index(columns.get("name"), "产品名称", sheet.title)
        status_idx = column_index(columns.get("status"), "产品状态", sheet.title)
        created_idx = column_index(columns.get("created"), "创建时间", sheet.title)
        department_idx = column_index(columns.get("department"), "部门", sheet.title)
        approval_idx = columns.get("delisting_approval_completed", {}).get("index") if columns.get("delisting_approval_completed") else None
        max_col = max(
            code_idx,
            name_idx,
            status_idx,
            created_idx,
            department_idx,
            int(approval_idx) if approval_idx is not None else 0,
        ) + 1
        sheet_count = 0
        for row_index, row in enumerate(sheet.iter_rows(min_row=sheet_mapping["data_start_row"], max_col=max_col, values_only=True), start=sheet_mapping["data_start_row"]):
            code = normalize(get_cell(row, code_idx))
            name = normalize(get_cell(row, name_idx))
            if not code and not name:
                continue
            if not code:
                raise ValueError(f"{sheet.title} 第 {row_index} 行缺少产品编码，无法按产品编码匹配")
            if code in products:
                raise ValueError(f"全量产品清单存在重复产品编码：{code}")
            status = normalize(get_cell(row, status_idx))
            created = get_cell(row, created_idx)
            approval_completed = get_cell(row, int(approval_idx) if approval_idx is not None else None)
            products[code] = {
                "产品编码": code,
                "产品名称": name,
                "产品状态": status,
                "创建时间": normalize(created),
                "创建日期": parse_date(created),
                "部门": normalize(get_cell(row, department_idx)),
                "退市审批完成时间": normalize(approval_completed),
                "退市审批完成日期": parse_date(approval_completed),
                "来源sheet": sheet.title,
                "来源行号": row_index,
            }
            status_counter[status] += 1
            sheet_count += 1
        sheet_counts[sheet.title] = sheet_count

    return products, {
        "product_count": len(products),
        "status_counts": dict(status_counter),
        "sheet_counts": sheet_counts,
    }


def read_revenue(revenue_paths: list[Path]) -> tuple[dict[str, dict[str, Any]], set[str], list[dict[str, Any]]]:
    aggregates = defaultdict(lambda: {"产品收入": Decimal("0"), "产品毛利": Decimal("0"), "all_rows": 0, "window_rows": 0})
    all_revenue_codes: set[str] = set()
    workbook_summaries: list[dict[str, Any]] = []

    for workbook_path in revenue_paths:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        if TARGET_REVENUE_SHEET not in workbook.sheetnames:
            raise ValueError(f"{workbook_path.name} 缺少 sheet：{TARGET_REVENUE_SHEET}")
        sheet = workbook[TARGET_REVENUE_SHEET]
        headers = [normalize(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        code_idx = first_existing(headers, ["产品编码"])
        revenue_idx = first_existing(headers, ["产品收入", "金额"])
        gross_profit_idx = first_existing(headers, ["产品毛利"])
        month_idx = first_existing(headers, ["收入月份"])
        missing = [
            label
            for label, idx in {
                "产品编码": code_idx,
                "产品收入/金额": revenue_idx,
                "产品毛利": gross_profit_idx,
                "收入月份": month_idx,
            }.items()
            if idx is None
        ]
        if missing:
            raise ValueError(f"{workbook_path.name} 缺少必要字段：{', '.join(missing)}")

        source_rows = 0
        window_rows = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            source_rows += 1
            code = normalize(row[code_idx])
            if not code:
                continue
            all_revenue_codes.add(code)
            aggregates[code]["all_rows"] += 1
            current_month = month_key(row[month_idx])
            if REVENUE_START_MONTH <= current_month <= REVENUE_END_MONTH:
                window_rows += 1
                aggregates[code]["window_rows"] += 1
                aggregates[code]["产品收入"] += to_decimal(row[revenue_idx])
                aggregates[code]["产品毛利"] += to_decimal(row[gross_profit_idx])

        workbook_summaries.append(
            {
                "file": workbook_path.name,
                "source_rows": source_rows,
                "window_rows": window_rows,
                "revenue_column": headers[revenue_idx],
            }
        )
    return aggregates, all_revenue_codes, workbook_summaries


def read_revenue_from_mapping(inspection: dict[str, Any]) -> tuple[dict[str, dict[str, Any]], set[str], list[dict[str, Any]]]:
    files_by_id = {item["id"]: item for item in inspection["files"]}
    aggregates = defaultdict(lambda: {"产品收入": Decimal("0"), "产品毛利": Decimal("0"), "all_rows": 0, "window_rows": 0})
    all_revenue_codes: set[str] = set()
    workbook_summaries: list[dict[str, Any]] = []

    for sheet_mapping in inspection["selected"]["revenue_sheets"]:
        file_info = files_by_id[sheet_mapping["file_id"]]
        workbook = load_workbook(file_info["path"], read_only=True, data_only=True)
        sheet = workbook[sheet_mapping["sheet"]]
        columns = sheet_mapping["columns"]
        code_idx = column_index(columns.get("code"), "产品编码", sheet.title)
        revenue_idx = column_index(columns.get("revenue"), "产品收入", sheet.title)
        gross_profit_idx = column_index(columns.get("gross_profit"), "产品毛利", sheet.title)
        month_idx = column_index(columns.get("month"), "收入月份", sheet.title)
        max_col = max(code_idx, revenue_idx, gross_profit_idx, month_idx) + 1
        source_rows = 0
        window_rows = 0
        for row in sheet.iter_rows(min_row=sheet_mapping["data_start_row"], max_col=max_col, values_only=True):
            source_rows += 1
            code = normalize(get_cell(row, code_idx))
            if not code:
                continue
            all_revenue_codes.add(code)
            aggregates[code]["all_rows"] += 1
            current_month = month_key(get_cell(row, month_idx))
            if REVENUE_START_MONTH <= current_month <= REVENUE_END_MONTH:
                window_rows += 1
                aggregates[code]["window_rows"] += 1
                aggregates[code]["产品收入"] += to_decimal(get_cell(row, revenue_idx))
                aggregates[code]["产品毛利"] += to_decimal(get_cell(row, gross_profit_idx))
        workbook_summaries.append(
            {
                "file": file_info["name"],
                "sheet": sheet.title,
                "source_rows": source_rows,
                "window_rows": window_rows,
                "revenue_column": columns["revenue"]["header"],
                "gross_profit_column": columns["gross_profit"]["header"],
                "month_column": columns["month"]["header"],
            }
        )

    return aggregates, all_revenue_codes, workbook_summaries


def product_base_row(product: dict[str, Any]) -> dict[str, Any]:
    return {
        "产品编码": product["产品编码"],
        "产品名称": product["产品名称"],
        "产品状态": product["产品状态"],
        "创建时间": product["创建时间"],
        "部门": product["部门"],
        "来源sheet": product["来源sheet"],
        "来源行号": product["来源行号"],
    }


def build_candidates(
    products: dict[str, dict[str, Any]],
    revenue: dict[str, dict[str, Any]],
    all_revenue_codes: set[str],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    candidates: list[dict[str, Any]] = []
    missing_approval_rows: list[dict[str, Any]] = []
    for code, product in products.items():
        status = product["产品状态"]
        revenue_values = revenue[code]
        product_revenue = revenue_values["产品收入"]
        product_gross_profit = revenue_values["产品毛利"]
        old_or_blank = product["创建日期"] is None or product["创建日期"] < OLDER_THAN_TWO_YEARS_BEFORE
        approval_completed_date = product.get("退市审批完成日期")
        active = status in ACTIVE_STATUSES_FOR_RULES_1_TO_3
        rules: list[str] = []

        if active and code not in all_revenue_codes:
            rules.append("1")
        if active and old_or_blank and product_revenue == 0:
            rules.append("2")
        if active and old_or_blank and product_gross_profit <= 0:
            rules.append("3")
        if status == "退市中":
            if approval_completed_date is None:
                missing_approval_rows.append(
                    {
                        **product_base_row(product),
                        "产品收入": decimal_to_number(product_revenue),
                        "产品毛利": decimal_to_number(product_gross_profit),
                        "退市审批完成时间": product.get("退市审批完成时间", ""),
                        "问题": "缺少退市审批完成时间，无法判断是否退市中超过1年",
                    }
                )
            elif approval_completed_date < RULE4_APPROVAL_BEFORE:
                rules.append("4")
        if not rules:
            continue

        delisting_type = "强制退市" if {"1", "2"} & set(rules) else "建议退市"
        candidates.append(
            {
                **product_base_row(product),
                "产品收入": decimal_to_number(product_revenue),
                "产品毛利": decimal_to_number(product_gross_profit),
                "退市类型": delisting_type,
                "理由": "；".join(RULE_TEXT[rule] for rule in rules),
                "命中规则": ",".join(rules),
                "退市审批完成时间": product.get("退市审批完成时间", ""),
            }
        )

    type_rank = {"强制退市": 0, "建议退市": 1}
    candidates.sort(key=lambda row: (type_rank[row["退市类型"]], row["命中规则"], row["产品编码"]))
    missing_approval_rows.sort(key=lambda row: (row["来源sheet"], row["产品编码"]))
    return candidates, missing_approval_rows


def make_payload(product_list_path: Path, revenue_paths: list[Path]) -> dict[str, Any]:
    products, product_summary = read_product_list(product_list_path)
    revenue, all_revenue_codes, workbook_summaries = read_revenue(revenue_paths)
    audit_rows, missing_approval_rows = build_candidates(products, revenue, all_revenue_codes)
    rule_counts: Counter[str] = Counter()
    type_counts: Counter[str] = Counter()
    candidate_status_counts: Counter[str] = Counter()
    for row in audit_rows:
        type_counts[row["退市类型"]] += 1
        candidate_status_counts[row["产品状态"]] += 1
        for rule in row["命中规则"].split(","):
            rule_counts[rule] += 1

    return {
        "headers": OUTPUT_HEADERS,
        "metadata": {
            "as_of_date": AS_OF_DATE.isoformat(),
            "older_than_two_years_before": OLDER_THAN_TWO_YEARS_BEFORE.isoformat(),
            "rule4_approval_before": RULE4_APPROVAL_BEFORE.isoformat(),
            "revenue_window": {"start_month": REVENUE_START_MONTH, "end_month": REVENUE_END_MONTH},
            "active_statuses_for_rules_1_to_3": sorted(ACTIVE_STATUSES_FOR_RULES_1_TO_3),
            "product_summary": product_summary,
            "revenue_workbooks": workbook_summaries,
            "all_revenue_code_count": len(all_revenue_codes),
            "candidate_count": len(audit_rows),
            "missing_rule4_approval_count": len(missing_approval_rows),
            "delisting_type_counts": dict(type_counts),
            "candidate_status_counts": dict(candidate_status_counts),
            "rule_counts": dict(rule_counts),
        },
        "rows": [{header: row[header] for header in OUTPUT_HEADERS} for row in audit_rows],
        "missing_rule4_approval_rows": missing_approval_rows,
        "audit_rows": audit_rows,
    }


def make_payload_from_mapping(inspection: dict[str, Any]) -> dict[str, Any]:
    products, product_summary = read_product_list_from_mapping(inspection)
    revenue, all_revenue_codes, workbook_summaries = read_revenue_from_mapping(inspection)
    audit_rows, missing_approval_rows = build_candidates(products, revenue, all_revenue_codes)
    rule_counts: Counter[str] = Counter()
    type_counts: Counter[str] = Counter()
    candidate_status_counts: Counter[str] = Counter()
    for row in audit_rows:
        type_counts[row["退市类型"]] += 1
        candidate_status_counts[row["产品状态"]] += 1
        for rule in row["命中规则"].split(","):
            rule_counts[rule] += 1

    return {
        "headers": OUTPUT_HEADERS,
        "metadata": {
            "as_of_date": AS_OF_DATE.isoformat(),
            "older_than_two_years_before": OLDER_THAN_TWO_YEARS_BEFORE.isoformat(),
            "rule4_approval_before": RULE4_APPROVAL_BEFORE.isoformat(),
            "revenue_window": {"start_month": REVENUE_START_MONTH, "end_month": REVENUE_END_MONTH},
            "active_statuses_for_rules_1_to_3": sorted(ACTIVE_STATUSES_FOR_RULES_1_TO_3),
            "product_summary": product_summary,
            "revenue_workbooks": workbook_summaries,
            "all_revenue_code_count": len(all_revenue_codes),
            "candidate_count": len(audit_rows),
            "missing_rule4_approval_count": len(missing_approval_rows),
            "delisting_type_counts": dict(type_counts),
            "candidate_status_counts": dict(candidate_status_counts),
            "rule_counts": dict(rule_counts),
            "inspection_warnings": inspection.get("warnings", []),
        },
        "rows": [{header: row[header] for header in OUTPUT_HEADERS} for row in audit_rows],
        "missing_rule4_approval_rows": missing_approval_rows,
        "audit_rows": audit_rows,
    }


def add_table(sheet, ref: str, name: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    sheet.add_table(table)


def write_output_workbook(payload: dict[str, Any], output_path: Path) -> None:
    workbook = Workbook()
    result = workbook.active
    result.title = "退市筛选结果"
    result.append(OUTPUT_HEADERS)
    for row in payload["rows"]:
        result.append([row.get(header, "") for header in OUTPUT_HEADERS])
    widths = [18, 34, 12, 22, 22, 16, 16, 14, 82]
    for index, width in enumerate(widths, start=1):
        result.column_dimensions[result.cell(1, index).column_letter].width = width
    for cell in result[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for row in result.iter_rows(min_row=2, min_col=6, max_col=7):
        for cell in row:
            cell.number_format = "#,##0.00"
    for row in result.iter_rows(min_row=2, min_col=8, max_col=9):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    result.freeze_panes = "A2"
    result.auto_filter.ref = result.dimensions
    add_table(result, result.dimensions, "DelistingScreeningTable")

    summary = workbook.create_sheet("口径说明")
    metadata = payload["metadata"]
    summary_rows = [
        ["项目", "说明"],
        ["基准日期", metadata["as_of_date"]],
        ["2年创建时间阈值", f"早于 {metadata['older_than_two_years_before']}，或创建时间为空"],
        ["规则4审批时间阈值", f"早于 {metadata['rule4_approval_before']}，即退市审批完成时间超过1年"],
        ["2年经营数据窗口", f"{metadata['revenue_window']['start_month']} 至 {metadata['revenue_window']['end_month']}"],
        ["规则1/2/3状态范围", "、".join(metadata["active_statuses_for_rules_1_to_3"])],
        ["规则4状态范围", "退市中；且退市审批完成时间超过1年"],
        ["候选总数", metadata["candidate_count"]],
        ["退市中缺少审批完成时间数量", metadata.get("missing_rule4_approval_count", 0)],
        ["强制退市数量", metadata["delisting_type_counts"].get("强制退市", 0)],
        ["建议退市数量", metadata["delisting_type_counts"].get("建议退市", 0)],
        ["规则1命中数", metadata["rule_counts"].get("1", 0)],
        ["规则2命中数", metadata["rule_counts"].get("2", 0)],
        ["规则3命中数", metadata["rule_counts"].get("3", 0)],
        ["规则4命中数", metadata["rule_counts"].get("4", 0)],
        ["全量产品数量", metadata["product_summary"]["product_count"]],
        ["全量产品状态分布", "；".join(f"{k}:{v}" for k, v in metadata["product_summary"]["status_counts"].items())],
        ["收入明细文件", "；".join(f"{item['file']}（{item['revenue_column']}，窗口内{item['window_rows']}行）" for item in metadata["revenue_workbooks"])],
    ]
    for row in summary_rows:
        summary.append(row)
    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 120
    for cell in summary[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
    for row in summary.iter_rows(min_row=2, min_col=2, max_col=2):
        row[0].alignment = Alignment(wrap_text=True, vertical="top")
    summary.freeze_panes = "A2"
    add_table(summary, summary.dimensions, "ScreeningBasisTable")

    rules = workbook.create_sheet("规则汇总")
    rule_rows = [
        ["规则", "命中数量", "退市类型影响", "说明"],
        ["规则1", metadata["rule_counts"].get("1", 0), "强制退市", RULE_TEXT["1"]],
        ["规则2", metadata["rule_counts"].get("2", 0), "强制退市", RULE_TEXT["2"]],
        ["规则3", metadata["rule_counts"].get("3", 0), "建议退市", RULE_TEXT["3"]],
        ["规则4", metadata["rule_counts"].get("4", 0), "建议退市", RULE_TEXT["4"]],
    ]
    for row in rule_rows:
        rules.append(row)
    for col, width in zip("ABCD", [12, 12, 16, 72]):
        rules.column_dimensions[col].width = width
    for cell in rules[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
    rules.freeze_panes = "A2"
    add_table(rules, rules.dimensions, "RuleSummaryTable")

    missing_approval_rows = payload.get("missing_rule4_approval_rows", [])
    missing_sheet = workbook.create_sheet("退市中缺审批时间")
    missing_headers = [
        "产品编码",
        "产品名称",
        "产品状态",
        "创建时间",
        "部门",
        "产品收入",
        "产品毛利",
        "退市审批完成时间",
        "问题",
        "来源sheet",
        "来源行号",
    ]
    missing_sheet.append(missing_headers)
    for row in missing_approval_rows:
        missing_sheet.append([row.get(header, "") for header in missing_headers])
    for col, width in zip("ABCDEFGHIJK", [18, 34, 12, 22, 22, 16, 16, 22, 46, 16, 12]):
        missing_sheet.column_dimensions[col].width = width
    for cell in missing_sheet[1]:
        cell.fill = PatternFill("solid", fgColor="9A3412")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for row in missing_sheet.iter_rows(min_row=2, min_col=6, max_col=7):
        for cell in row:
            cell.number_format = "#,##0.00"
    for row in missing_sheet.iter_rows(min_row=2, min_col=8, max_col=9):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    missing_sheet.freeze_panes = "A2"
    add_table(missing_sheet, missing_sheet.dimensions, "MissingRule4ApprovalTable")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--inspect-files", nargs="+")
    parser.add_argument("--mapping-json")
    parser.add_argument("--product-list")
    parser.add_argument("--revenue-files", nargs="+")
    parser.add_argument("--output-json", required=True)
    parser.add_argument("--output-xlsx")
    args = parser.parse_args()

    output_json = Path(args.output_json)
    output_json.parent.mkdir(parents=True, exist_ok=True)

    if args.inspect_files:
        inspection = inspect_workbooks([Path(item) for item in args.inspect_files])
        output_json.write_text(json.dumps(inspection, ensure_ascii=False, indent=2), encoding="utf-8")
        print(json.dumps({"ok": True, "mode": "inspect", "files": len(inspection["files"])}, ensure_ascii=False))
        return

    if args.mapping_json:
        inspection = json.loads(Path(args.mapping_json).read_text(encoding="utf-8"))
        payload = make_payload_from_mapping(inspection)
    else:
        if not args.product_list or not args.revenue_files:
            raise ValueError("缺少 --product-list/--revenue-files 或 --mapping-json")
        payload = make_payload(Path(args.product_list), [Path(item) for item in args.revenue_files])

    output_json.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    if args.output_xlsx:
        write_output_workbook(payload, Path(args.output_xlsx))
    print(json.dumps({"ok": True, "metadata": payload["metadata"]}, ensure_ascii=False))


if __name__ == "__main__":
    main()
